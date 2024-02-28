import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the existing workbook
workbook = load_workbook('existing_file.xlsx')

# Select the worksheet where you want to insert the DataFrame
worksheet = workbook['Sheet1']  # Change 'Sheet1' to your sheet name

# Create or load your DataFrame
data = {'Column1': [1, 2, 3],
        'Column2': ['A', 'B', 'C']}
df = pd.DataFrame(data)

# Get the rows from the DataFrame
rows = dataframe_to_rows(df, index=False, header=True)

# Determine the starting cell for the DataFrame
start_row = 1  # You can adjust this based on where you want to insert the DataFrame
start_column = 1

# Write the DataFrame to the worksheet
for r_idx, row in enumerate(rows, start=start_row):
    for c_idx, value in enumerate(row, start=start_column):
        worksheet.cell(row=r_idx, column=c_idx, value=value)

# Save the changes
workbook.save('existing_file.xlsx')




from openpyxl import load_workbook

# Load the existing workbook
workbook = load_workbook('existing_file.xlsx')

# Select the worksheet where you want to search for the lowest row with a value
worksheet = workbook['Sheet1']  # Change 'Sheet1' to your sheet name

# Specify the column to search for the lowest row with a value
column_to_search = 'A'  # Change 'A' to the column you want to search

# Iterate through the rows in the specified column
lowest_row_with_value = None
for row in worksheet.iter_rows(min_row=1, min_col=column_to_search, max_col=column_to_search, values_only=True):
    for cell_value in row:
        if cell_value is not None:
            # Found the first non-empty cell in the column
            lowest_row_with_value = row[0].row
            break
    if lowest_row_with_value is not None:
        break

if lowest_row_with_value is not None:
    print(f"The lowest row with a value in column {column_to_search} is: {lowest_row_with_value}")
else:
    print(f"No value found in column {column_to_search}")
    
    
from openpyxl import load_workbook

# Load the existing workbook
workbook = load_workbook('existing_file.xlsx')

# Select the worksheet where you want to delete rows
worksheet = workbook['Sheet1']  # Change 'Sheet1' to your sheet name

# Specify the row number from which you want to start deleting rows
start_row_to_delete = 10  # Change 10 to the row number you want to start deleting from

# Delete rows below the specified row
max_row = worksheet.max_row
for row in range(start_row_to_delete + 1, max_row + 1):
    worksheet.delete_rows(start_row_to_delete + 1)

# Save the changes
workbook.save('existing_file.xlsx')